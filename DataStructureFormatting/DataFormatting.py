from openpyxl import load_workbook
from PIL import Image
import os
import imageio
import numpy as np


def generateStructures(SSfilename, IMdirectory, maxRow):
    #we need to return a list of the image files themselves
    #we need to return a list of the characteristic data

    wb = load_workbook(SSfilename) #Load Spreadsheet
    ws = wb.active #Grab Active Sheet

    os.chdir(IMdirectory) #Change directory to where images are
    filenames = [i for i in ws.iter_cols(min_row=2,max_row=maxRow,min_col=1,max_col=1,values_only=True)][0] #grab image filenames
    condition = [i for i in ws.iter_cols(min_row=2,max_row=maxRow,min_col=2,max_col=2,values_only=True)][0] #get eye labels
    #images = [((imageio.imread(i))/255) for i in filenames] #open image data as an rgb matrix
    images = [Image.open(i) for i in filenames]
    #print(filenames)
    #print(condition)
    #print(images)

    return images, condition


if __name__ == "__main__":
    f,n = generateStructures(r"G:\Desktop\DataSets\Useful Datasets\FinalSet\testingData.xlsx",r"G:\Desktop\DataSets\Useful Datasets\FinalSet\Cropped",5)
